import argparse
import calendar
import csv
import glob
import math
import os
import pathlib
import sys
import unicodedata
import shutil
from datetime import date, timedelta, datetime
from dateutil.relativedelta import relativedelta

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

from s3_functions import *


class SimilarDayReport:
    companies = {'CKY': 32, 'COH': 34, 'CMD': 35, 'CPA': 37, 'CGV': 38, 'CMA': 80}

    def __init__(self, daily_file=None, report_file=None):
        self.bucket = 'isg-nisource-test-bucket'
        # self.list_files =
        self.archive_dir = 'historical'
        self.backup_dir = 'bak'

        self.report_filepath = self.get_current_file(self.bucket, 'xlsx')
        self.report_daily_file = self.get_current_file(self.bucket, '.csv')

        #         self.hist_dir = pathlib.Path(base_dir, 'historical')
        #         self.archive_dir = pathlib.Path(base_dir, 'archive')
        #         self.backup_dir = pathlib.Path(base_dir, 'bak')

        #         self.wb = load_workbook(self.report_filepath, data_only=True)
        file = download_file(self.bucket, self.report_filepath)
        self.wb = load_workbook(file)

    def generate(self, num_matches, save=False, overwrite=False, logging=False):
        #         print(f'\nRunning report generation for {self.report_filepath} with daily data: {self.daily_filepath}')

        #         if save:
        #             self.create_backup()

        df_daily = self.load_daily(self.bucket, self.report_daily_file)
        for sheet in self.wb.worksheets:
            #             hist_path = pathlib.Path(self.bucket, f'{sheet.title} data.csv')
            df_hist = self.load_historical(self.bucket, f'{sheet.title} data.csv')
            df_company = df_daily[df_daily['COMPANY'] == self.companies[sheet.title]]
            for cell in sheet.iter_cols(min_col=2, min_row=4, max_row=4):
                col = get_column_letter(cell[0].column)
                report_dt = sheet[f'{col}4'].value

                # if current column is in daily data and it's not already filled in
                if report_dt in df_company['GAS_DATE'].to_list():
                    has_empty_values = sheet[f'{col}6'].value is None or sheet[f'{col}16'].value is None or sheet[
                        f'{col}23'].value is None or sheet[f'{col}30'].value is None
                    if has_empty_values or overwrite:
                        print(f'Found blank day in {sheet.title}: ', sheet[f'{col}4'].value)
                        df_day = df_company[df_company['GAS_DATE'] == report_dt]
                        df_matches = self.find_similar(df_day, df_hist, num_matches)

                        avg_similar_day = (df_matches.iloc[0]['DTH'] + df_matches.iloc[1]['DTH'] + df_matches.iloc[2][
                            'DTH']) / 3
                        pct_diff = ((df_day.iloc[0]['DTH'] - avg_similar_day) / avg_similar_day)

                        if logging:
                            self.pprint(df_day, pct_diff, avg_similar_day, df_matches)

                        sheet[f'{col}12'] = pct_diff
                        sheet[f'{col}14'] = avg_similar_day

                        sheet[f'{col}{6}'] = df_day.iloc[0]['DTH']
                        sheet[f'{col}{8}'] = df_day.iloc[0]['GAS_DAY_AVG_TMP']
                        sheet[f'{col}{9}'] = df_day.iloc[0]['PRIOR_TEMP']
                        sheet[f'{col}{10}'] = df_day.iloc[0]['GAS_DAY_WIND_SPEED']

                        for i in range(0, len(df_matches)):
                            sheet[f'{col}{16 + (7 * i)}'] = df_matches.iloc[i]['DTH']
                            sheet[f'{col}{17 + (7 * i)}'] = df_matches.iloc[i]['GAS_DATE']
                            sheet[f'{col}{18 + (7 * i)}'] = df_matches.iloc[i]['DAY_SHORTNAME']
                            sheet[f'{col}{19 + (7 * i)}'] = df_matches.iloc[i]['GAS_DAY_AVG_TMP']
                            sheet[f'{col}{20 + (7 * i)}'] = df_matches.iloc[i]['PRIOR_TEMP']
                            sheet[f'{col}{21 + (7 * i)}'] = df_matches.iloc[i]['GAS_DAY_WIND_SPEED']

    #         if save:
    #             self.save()
    #             self.delete_backup()

    def get_current_file(self, bucket, suffix):
        s3 = boto3.client('s3')
        response = s3.list_objects(Bucket=bucket)

        # response['Contents']
        file_dict = {}

        for file in response['Contents']:
            if file['Key'][-4:] == suffix:
                file_dict[file['Key']] = file['LastModified']

        return max(file_dict, key=lambda k: file_dict[k] if isinstance(file_dict[k], datetime) else datetime.max)

    def load_daily(self, bucket, key):
        df = read_csv(bucket, key)
        df.rename(columns={'DayType': 'DAY_TYPE'}, inplace=True)  # rename for column name consistency
        df['DTH'] = df['DTH'].apply(lambda x: x / 1000)  # convert to dekatherm
        df['GAS_DATE'] = pd.to_datetime(df['GAS_DATE'])  # convert to datetime
        df['DAY_SHORTNAME'] = df['GAS_DATE'].dt.dayofweek.apply(to_dayname)  # add shortname column (Mon, Tues)
        df.sort_values(by=['COMPANY']).reset_index()
        return df

    def load_historical(self, bucket, key):
        df = read_csv(bucket, key)
        df.rename(columns={'DayType': 'DAY_TYPE'}, inplace=True)  # rename for column name consistency
        df['DTH'] = df['DTH'].apply(lambda x: x / 1000)  # convert to dekatherm
        df['GAS_DATE'] = pd.to_datetime(df['GAS_DATE'])  # convert to datetime
        df['DAY_SHORTNAME'] = df['GAS_DATE'].dt.dayofweek.apply(to_dayname)  # add shortname column (Mon, Tues)
        df.reset_index()
        return df

    def find_similar(self, df_day, df_hist, num_matches):
        """Criteria:
        +/- 2 degrees
        Start on minus year, same day"""
        factor_year = 1.0
        factor_dayofyear = 1.0
        factor_time = 10.0
        factor_wind = 1.25
        factor_dayofweek = 3.0

        end_range = df_day.iloc[0]['GAS_DATE'] + relativedelta(months=-1)
        start_range = df_day.iloc[0]['GAS_DATE'] + relativedelta(months=-25)
        print(f'Using date range {start_range} => {end_range.dt}')

        df_work = df_hist.copy(deep=True)

        df_work = df_work[df_work['GAS_DATE'] > pd.to_datetime(start_range)]
        df_work = df_work[df_work['GAS_DATE'] < pd.to_datetime(end_range)]

        df_work = df_work[df_work['GAS_DAY_AVG_TMP'] > df_day.iloc[0]['GAS_DAY_AVG_TMP'] - 2]
        df_work = df_work[df_work['GAS_DAY_AVG_TMP'] < df_day.iloc[0]['GAS_DAY_AVG_TMP'] + 2]

        df_work['YEAR_DELTA'] = (abs(df_work['GAS_DATE'].dt.year - df_day.iloc[0]['GAS_DATE'].year) + 1) * factor_year
        df_work['DAYOFYEAR_DELTA'] = (abs(
            df_work['GAS_DATE'].dt.dayofyear - df_day.iloc[0]['GAS_DATE'].dayofyear) + 1) * factor_dayofyear
        df_work['DAYOFWEEK_MULTIPLE'] = (abs(
            (df_work['DAY_TYPE'] == df_day.iloc[0]['DAY_TYPE']).astype(int) - 1) + 1) ** factor_dayofweek
        df_work['WIND_DELTA'] = abs(df_work['GAS_DAY_WIND_SPEED'] - df_day.iloc[0]['GAS_DAY_WIND_SPEED']) * factor_wind

        df_work['TMP_DELTA'] = abs(
            df_work['GAS_DAY_AVG_TMP'] - df_day.iloc[0]['GAS_DAY_AVG_TMP'])  # display only, not used in final weight
        df_work['DTH_DELTA'] = abs(df_work['DTH'] - df_day.iloc[0]['DTH'])  # display only, not used in final weight

        df_work['TIME_DELTA'] = (df_work['YEAR_DELTA'] + df_work['DAYOFYEAR_DELTA']) * factor_time
        df_work['DELTA_WEIGHTED'] = (df_work['TIME_DELTA'] - df_work['WIND_DELTA']) * df_work['DAYOFWEEK_MULTIPLE']

        df_work.sort_values(by=['DELTA_WEIGHTED'], inplace=True)

        return df_work.head(num_matches).reset_index()

    def save(self, key):
        #         # Save old workbook with current date inserted (move from /bak to /archive and add date)
        #         report_bakfile = (self.report_filepath.parent / self.report_filepath.name).with_suffix(
        #             self.report_filepath.suffix + '.bak').parts[-1]
        # #         report_bak = pathlib.Path(pathlib.Path.joinpath(self.backup_dir, report_bakfile))
        #         report_bak.replace(pathlib.Path.joinpath(self.archive_dir, 'reports', with_date(self.report_filepath)))

        # Save new workbook
        #         self.wb.save(self.report_filepath)
        upload_workbook(self.wb, self.bucket, key)

        # # Archive daily data
        #         daily_bakfile = \
        #         (self.daily_filepath.parent / self.daily_filepath.name).with_suffix(self.daily_filepath.suffix + '.bak').parts[
        #             -1]
        # #         daily_bak = pathlib.Path(pathlib.Path.joinpath(self.backup_dir, daily_bakfile))
        #         daily_bak.replace(pathlib.Path.joinpath(self.archive_dir, 'daily', self.daily_filepath.parts[-1]))

        # Append daily data to each company file
        df_daily = read_csv(self.bucket, key)
        for company_code in df_daily['COMPANY'].unique():
            company_name = list(self.companies.keys())[list(self.companies.values()).index(company_code)]
            key = f'{company_name} data.csv'
            #             company_filepath = pathlib.Path.joinpath(self.hist_dir, f'{company_name} data.csv')
            df_hist = read_csv(self.bucket, key)

            # append, making sure not to re-insert if data already present
            df = pd.concat([df_hist, df_daily[df_daily['COMPANY'] == company_code]]).drop_duplicates().sort_values(
                by='GAS_DATE').reset_index(drop=True)
            #             df.to_csv(company_filepath, index=False)
            write_csv(df, self.bucket, key)

        return True

    #     def create_backup(self):
    #         """Creates temporary backups for all files in case rollback is needed.
    #         These backups are deleted upon successful save.
    #         """
    #         pathlib.Path(self.backup_dir).mkdir(exist_ok=True)  # create backup directory, overwriting if necessary
    #         pathlib.Path(pathlib.Path.joinpath(self.backup_dir, self.hist_dir.parts[-1])).mkdir(
    #             exist_ok=True)  # create history directory

    #         # Create backups
    #         daily_bakfile = \
    #         (self.daily_filepath.parent / self.daily_filepath.name).with_suffix(self.daily_filepath.suffix + '.bak').parts[
    #             -1]
    #         daily_bak = shutil.copy(self.daily_filepath, pathlib.Path.joinpath(self.backup_dir, daily_bakfile))
    #         print(f'{daily_bak} created')

    #         report_bakfile = (self.report_filepath.parent / self.report_filepath.name).with_suffix(
    #             self.report_filepath.suffix + '.bak').parts[-1]
    #         report_bak = shutil.copy(self.report_filepath, pathlib.Path.joinpath(self.backup_dir, report_bakfile))
    #         print(f'{report_bak} created')

    #         for f in self.hist_dir.iterdir():
    #             hist_bakfile = (f.parent / f.name).with_suffix(f.suffix + '.bak').parts[-1]
    #             hist_bak = shutil.copy(f, pathlib.Path.joinpath(self.backup_dir, self.hist_dir.parts[-1], hist_bakfile))
    #             print(f'{hist_bak} created')

    #         # validation check
    #         daily_ok = daily_bak.is_file()
    #         report_ok = report_bak.is_file()
    #         hist_ok = all(f.is_file() for f in self.hist_dir.iterdir())

    #         if not (daily_ok and report_ok and hist_ok):
    #             print('Backup creation failed. Exiting...')
    #             sys.exit(1)

    #     def delete_backup(self):
    #         for f in self.backup_dir.rglob('*'):
    #             shutil.rmtree(f)
    #         self.backup_dir.rmdir()
    #         print('Successfully deleted backups')
    #         return True

    def pprint(self, df_day, pct_diff, avg_similar_day, df_matches):
        print(f'{df_day}')
        print(f'{df_matches}\n')
        print(f'Percent difference: {pct_diff * 100:.2f}%')
        print(f'Avg similar days: {avg_similar_day}\n')
        print(f'{"-" * 80}\n')

    def log(self, *args):
        for arg in args:
            if type(arg) == list:
                for a in arg:
                    print(f'{a}')
                print()
            else:
                print(f'{arg}\n')


def to_dayname(day):
    return ['Mon', 'Tues', 'Wed', 'Thurs', 'Fri', 'Sat', 'Sun'][day]


def to_daytype(day):
    return 'Weekend' if day.upper() in ['SAT', 'SUN'] else 'Weekday'

# def with_date(filename: pathlib.Path):
#     """Takes in a filename and returns it with a date inserted
#     Example:
#     Similar Reports April.xlsx => Similar Reports Aprils_20200101.xlsx
#     """
#     stem = filename.stem
#     dt = (date.today()-timedelta(1)).strftime('%Y%m%d') # yesterday, in format 20200101
#     sfxs = ''.join(filename.suffix)
#     return pathlib.Path(f'{stem}_{dt}{sfxs}')