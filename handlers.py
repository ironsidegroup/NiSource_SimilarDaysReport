import os
import sys
import re
import io

import boto3
import dropbox
from dropbox.exceptions import AuthError
from dropbox.files import WriteMode

import argparse
import calendar
import csv
import glob
import math
import os
import pathlib
import re
import shutil
import sys
import unicodedata
from datetime import date, timedelta
from tempfile import NamedTemporaryFile

import boto3
import botocore
import dropbox
import pandas as pd
from dateutil.relativedelta import relativedelta
from dropbox.exceptions import AuthError
from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter


class DropboxHandler:

    def __init__(self, token=None):
        self.client = self.connect(token)
        self.dropbox_folder = '/NiSource/'

    def connect(self, token=None):
        if not token:
            token = os.getenv('DROPBOX_ACCESS_TOKEN')
        if len(token) == 0 or token is None:
            sys.exit('Please ensure the Dropbox access token has been added to the AWS environment')

        print('Creating a Dropbox object...')
        dbx = dropbox.Dropbox(token)

        try:
            dbx.users_get_current_account()
            print(f'Valid connection')
        except AuthError:
            sys.exit(
                'ERROR: Invalid Dropbox access token; try re-generating an access token from the app console on the web.')

        return dbx

    def get_newest_file(self, pattern):
        res = []
        for e in self.client.files_list_folder(self.dropbox_folder).entries:

            if re.match(pattern, e.name) and type(e) == dropbox.files.FileMetadata:
                res.append(e)

        try:
            newest = max(res, key=lambda e: e.server_modified)
        except ValueError as e:
            print(e)
            # os.exit(1)
            quit()

        metadata, f = self.client.files_download(newest.path_lower)
        return metadata, f

    def get_files(self, dir):
        res = []
        for e in self.client.files_list_folder(self.dropbox_folder + dir).entries:
            if type(e) == dropbox.files.FileMetadata:
                metadata, f = self.client.files_download(e.path_lower)
                res.append((metadata, f))
        return res


class S3Handler:

    def __init__(self, bucket):
        self.client = boto3.client('s3')
        self.resource = boto3.resource('s3')
        self.bucket = self.resource.Bucket(bucket)
        self.dropbox = DropboxHandler()

        self.tmp_dir = '/tmp/'
        self.hist_dir = 'historical/'

    def stage(self):
        """Pulls the newest daily and report files.
        Pulls the 6 historical company files.
        Writes all files to S3 bucket.
        Returns a dict with all filenames."""

        daily_metadata, daily_file = self.dropbox.get_newest_file(r'^[0-9]{1,2}-[0-9]{1,2} data\.csv$')
        report_metadata, report_file = self.dropbox.get_newest_file(r'^Similar Days \w+\.xlsx$')
        hist_files = self.dropbox.get_files(self.hist_dir)

        # Put files in bucket
        self.bucket.put_object(Body=daily_file.content, Key=daily_metadata.name)
        self.bucket.put_object(Body=report_file.content, Key=report_metadata.name)
        for hist_metadata, hist_file in hist_files:
            self.bucket.put_object(Body=hist_file.content, Key=self.hist_dir + hist_metadata.name)

        filenames = {}
        filenames['daily'] = self.download_csv(daily_metadata.name)
        filenames['report'] = self.download_excel(report_metadata.name)
        filenames['hist'] = [self.download_csv(md.name, self.hist_dir) for (md, h) in hist_files]
        print(filenames)
        return filenames

    def unstage(self, report, filenames):
        df_daily = pd.read_csv(filenames['daily'])
        for company_code in df_daily['COMPANY'].unique():
            company_name = list(ReportHandler.companies.keys())[
                list(ReportHandler.companies.values()).index(company_code)]
            hist_file = pathlib.Path(self.tmp_dir, self.hist_dir, f'{company_name} data.csv')
            print(f'Appending {company_name} data to {str(hist_file)}')
            df_hist = pd.read_csv(hist_file)

            # append, making sure not to re-insert if data already present
            df = pd.concat([df_hist, df_daily[df_daily['COMPANY'] == company_code]]).drop_duplicates().sort_values(
                by='GAS_DATE').reset_index(drop=True)
            print(f'Saving {hist_file}')
            df.to_csv(hist_file, index=False)

        # Archive
        print(f'Archiving {filenames["daily"].name}')
        print(f'Archiving {filenames["report"].name}')
        self.dropbox.client.files_move_v2(self.dropbox.dropbox_folder + filenames['daily'].name,
                                          self.dropbox.dropbox_folder + 'archive/daily/' + filenames['daily'].name,
                                          autorename=True)
        self.dropbox.client.files_move_v2(self.dropbox.dropbox_folder + filenames['report'].name,
                                          self.dropbox.dropbox_folder + 'archive/reports/' + filenames['report'].name,
                                          autorename=True)

        today_report_name = with_date(filenames['report']).name
        report.save(self.tmp_dir + today_report_name)
        print(f'Uploading {filenames["report"]} to Dropbox')
        with open(self.tmp_dir + today_report_name, 'rb') as f:
            self.dropbox.client.files_upload(f.read(), self.dropbox.dropbox_folder + today_report_name)

        for hist_file in filenames['hist']:
            print(f'Uploading {hist_file} to Dropbox')
            with open(hist_file, 'rb') as f:
                self.dropbox.client.files_upload(f.read(), self.dropbox.dropbox_folder + self.hist_dir + hist_file.name,
                                                 mode=WriteMode('overwrite'))

    def download_excel(self, key, dir=''):
        try:
            if dir:
                pathlib.Path(self.tmp_dir, dir).mkdir(exist_ok=True)

            tmp_path = pathlib.Path(self.tmp_dir, dir, key)
            print(f'Writing file to {tmp_path}')
            self.bucket.download_file(key, str(tmp_path))
            return tmp_path
        except botocore.exceptions.ClientError as e:
            print('ERROR', e)
            if e.response['Error']['Code'] == '404':
                return None
            else:
                raise
        else:
            raise

    def download_csv(self, key, dir=''):
        try:
            if dir:
                pathlib.Path(self.tmp_dir, dir).mkdir(exist_ok=True)

            tmp_path = pathlib.Path(self.tmp_dir, dir, key)
            print(f'Writing file to {tmp_path}')
            self.bucket.download_file(key, str(tmp_path))
            return tmp_path
        except botocore.exceptions.ClientError as e:
            if e.response['Error']['Code'] == '404':
                return None
            else:
                raise
        else:
            raise


class ReportHandler:
    companies = {'CKY': 32, 'COH': 34, 'CMD': 35, 'CPA': 37, 'CGV': 38, 'CMA': 80}

    def __init__(self, report_file, daily_file, history_dir):
        self.base_dir = '/tmp/'
        self.daily_file = pathlib.Path(self.base_dir, daily_file)
        self.report_file = pathlib.Path(self.base_dir, report_file)
        self.hist_dir = pathlib.Path(self.base_dir, history_dir)

    def generate(self, num_matches, logging=False, overwrite=False):
        print(f'Generating report: {self.daily_file} => {self.report_file}')

        wb = load_workbook(self.report_file, data_only=False)

        df_daily = self.load_data(self.daily_file)
        for sheet in wb.worksheets:
            if sheet.title == 'Summary':
                continue
            hist_path = pathlib.Path(self.hist_dir, f'{sheet.title} data.csv')
            df_hist = self.load_data(hist_path)
            df_company = df_daily[df_daily['COMPANY'] == self.companies[sheet.title]]
            for cell in sheet.iter_cols(min_col=2, min_row=4, max_row=4):
                col = get_column_letter(cell[0].column)
                report_dt = sheet[f'{col}4'].value

                # if current column is in daily data and it's not already filled in
                if report_dt in df_company['GAS_DATE'].to_list():
                    has_empty_values = sheet[f'{col}6'].value is None or sheet[f'{col}16'].value is None or sheet[
                        f'{col}23'].value is None or sheet[f'{col}30'].value is None
                    if has_empty_values or overwrite:
                        print(f'Adding {sheet[f"{col}4"].value} to {sheet.title}')
                        df_day = df_company[df_company['GAS_DATE'] == report_dt]
                        df_matches = self.find_similar(df_day, df_hist, num_matches)

                        avg_similar_day = (df_matches.iloc[0]['DTH'] + df_matches.iloc[1]['DTH'] + df_matches.iloc[2][
                            'DTH']) / 3
                        pct_diff = ((df_day.iloc[0]['DTH'] - avg_similar_day) / avg_similar_day)

                        if logging:
                            self.pprint(df_day, pct_diff, avg_similar_day, df_matches)

                        # sheet[f'{col}12'] = pct_diff
                        # sheet[f'{col}14'] = avg_similar_day

                        sheet[f'{col}{6}'] = df_day.iloc[0]['DTH']
                        sheet[f'{col}{8}'] = df_day.iloc[0]['GAS_DAY_AVG_TMP']
                        sheet[f'{col}{9}'] = df_day.iloc[0]['PRIOR_TEMP']
                        sheet[f'{col}{10}'] = df_day.iloc[0]['GAS_DAY_WIND_SPEED']

                        for i in range(0, len(df_matches)):
                            sheet[f'{col}{17 + (7 * i)}'] = df_matches.iloc[i]['DTH']
                            sheet[f'{col}{18 + (7 * i)}'] = df_matches.iloc[i]['GAS_DATE']
                            sheet[f'{col}{19 + (7 * i)}'] = df_matches.iloc[i]['DAY_SHORTNAME']
                            sheet[f'{col}{20 + (7 * i)}'] = df_matches.iloc[i]['GAS_DAY_AVG_TMP']
                            sheet[f'{col}{21 + (7 * i)}'] = df_matches.iloc[i]['PRIOR_TEMP']
                            sheet[f'{col}{22 + (7 * i)}'] = df_matches.iloc[i]['GAS_DAY_WIND_SPEED']

        return wb

    def load_data(self, csv_file):
        df = pd.read_csv(csv_file)
        df.rename(columns={'DayType': 'DAY_TYPE'}, inplace=True)  # rename for column name consistency
        df['DTH'] = df['DTH'].apply(lambda x: x / 1000)  # convert to dekatherm
        df['GAS_DATE'] = pd.to_datetime(df['GAS_DATE'])  # convert to datetime
        df['DAY_SHORTNAME'] = df['GAS_DATE'].dt.dayofweek.apply(to_dayname)  # add shortname column (Mon, Tues)
        df.sort_values(by=['COMPANY']).reset_index()
        return df

    def find_similar(self, df_day, df_hist, num_matches):
        """Criteria:
        +/- 2 degrees
        # Start on minus year, same day"""
        factor_year = 0.10
        factor_dayofyear = 0.40
        factor_wind = 0.05
        factor_daytype = 0.45

        end_range = df_day.iloc[0]['GAS_DATE'] + relativedelta(months=-1)
        start_range = df_day.iloc[0]['GAS_DATE'] + relativedelta(months=-25)
        print(f'Using date range {start_range} => {end_range}')

        df_work = df_hist.copy(deep=True)

        df_work = df_work[df_work['GAS_DATE'] >= pd.to_datetime(start_range)]
        df_work = df_work[df_work['GAS_DATE'] <= pd.to_datetime(end_range)]

        df_work = df_work[df_work['GAS_DAY_AVG_TMP'] >= df_day.iloc[0]['GAS_DAY_AVG_TMP'] - 2]
        df_work = df_work[df_work['GAS_DAY_AVG_TMP'] <= df_day.iloc[0]['GAS_DAY_AVG_TMP'] + 2]

        df_work['TMP_DELTA'] = abs(
            df_work['GAS_DAY_AVG_TMP'] - df_day.iloc[0]['GAS_DAY_AVG_TMP'])  # display only, not used in final weight
        df_work['DTH_DELTA'] = abs(df_work['DTH'] - df_day.iloc[0]['DTH'])  # display only, not used in final weight

        df_work['YEAR_DELTA'] = abs(df_work['GAS_DATE'].dt.year - df_day.iloc[0]['GAS_DATE'].year) + 1
        df_work['DAYOFYEAR_DELTA'] = abs(df_work['GAS_DATE'].dt.dayofyear - df_day.iloc[0]['GAS_DATE'].dayofyear) + 1
        df_work['WIND_DELTA'] = abs(df_work['GAS_DAY_WIND_SPEED'] - df_day.iloc[0]['GAS_DAY_WIND_SPEED']) + 1
        df_work['DAYTYPE_DELTA'] = abs(
            (df_work['DAY_TYPE'] == df_day.iloc[0]['DAY_TYPE']).astype(int) - 1) + 1  # 0 if same day, 1 if not

        df_work['YEAR_FACTOR'] = df_work['YEAR_DELTA'] / df_work['GAS_DATE'].dt.year.mean()
        df_work['DAYOFYEAR_FACTOR'] = df_work['DAYOFYEAR_DELTA'] / df_work['GAS_DATE'].dt.dayofyear.mean()
        df_work['WIND_FACTOR'] = df_work['WIND_DELTA'] / df_work['GAS_DAY_WIND_SPEED'].mean()
        df_work['DAYTYPE_FACTOR'] = df_work['DAYTYPE_DELTA']

        df_work['WEIGHTED_FACTOR'] = \
            (df_work['YEAR_FACTOR'] * factor_year) \
            + (df_work['DAYOFYEAR_FACTOR'] * factor_dayofyear) \
            + (df_work['WIND_FACTOR'] * factor_wind) \
            + ((df_work['DAYTYPE_FACTOR'] * factor_daytype) / 2)

        df_work.sort_values(by=['WEIGHTED_FACTOR'], inplace=True)

        return df_work.head(num_matches).reset_index()

    def pprint(self, df_day, pct_diff, avg_similar_day, df_matches):
        print(f'{df_day}')
        print(f'{df_matches}\n')
        print(f'Percent difference: {pct_diff * 100:.2f}%')
        print(f'Avg similar days: {avg_similar_day}\n')
        print(f'{"-" * 80}\n')


def to_dayname(day):
    return ['Mon', 'Tues', 'Wed', 'Thurs', 'Fri', 'Sat', 'Sun'][day]


def to_daytype(day):
    return 'Weekend' if day.upper() in ['SAT', 'SUN'] else 'Weekday'


def with_date(filename: pathlib.Path, tdelta=0):
    """Takes in a filename and returns it with a date inserted
    Example:
    Similar Reports April.xlsx => Similar Reports Aprils_20200101.xlsx
    """
    date_pattern = r'[0-9]{4}[0-9]{2}[0-9]{2}'
    stem = filename.stem.split('_')[0] if re.search(date_pattern, filename.stem) is not None else filename.stem

    dt = (date.today() + timedelta(tdelta)).strftime('%Y%m%d')  # today, in format 20200101
    sfxs = ''.join(filename.suffix)
    return pathlib.Path(f'{stem}_{dt}{sfxs}')